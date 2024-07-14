import streamlit as st
import pandas as pd
import yfinance as yf
from ta.volatility import BollingerBands
from ta.trend import MACD, EMAIndicator, SMAIndicator
from ta.momentum import RSIIndicator
import datetime
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor
from xgboost import XGBRegressor
from sklearn.metrics import r2_score, mean_absolute_error
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from random import randint
from email_validator import validate_email, EmailNotValidError
from openpyxl import load_workbook
from datetime import datetime as dt
import os
import numpy as np

@st.cache
def load_animation_file(file_path):
    try:
        with open(file_path, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        st.error(f"Animation file not found: {file_path}")
        return None
st.set_page_config(page_title="team-tanx")
st.title('Stock Market Predictions Using Streamlit')

paragraph = '''
<p style="color:red;">
Predicting stock market prices carries significant risks due to the inherent volatility and complexity of financial markets. 
Factors such as economic indicators, geopolitical events, company performance, investor sentiment, and unexpected news can all 
cause sudden and unpredictable market fluctuations. Even advanced models using sophisticated algorithms and historical data can 
fail to account for these rapid changes, leading to inaccurate predictions. Additionally, reliance on technical indicators and 
past trends can be misleading in the face of market anomalies. Furthermore, overconfidence in predictions can lead to substantial 
financial losses, emphasizing the importance of caution and diversified investment strategies to mitigate potential risks.
</p>
'''

st.markdown(paragraph, unsafe_allow_html=True)

st.sidebar.info("Created and designed by Team TanX")

def download_data(stock_symbol, start_date, end_date):
    try:
        stock_info = yf.Ticker(stock_symbol)
        company_name = stock_info.info['longName']
        st.sidebar.write(f"Stock: {stock_symbol} - {company_name}")
        df = yf.download(stock_symbol, start=start_date, end=end_date, progress=False)
        return df
    except Exception as e:
        st.error(f"Error downloading data: {e}")
        return pd.DataFrame()

def tech_indicators(data):
    st.header('Technical Indicators')
    option = st.radio('Choose a Technical Indicator to Visualize', ['Close', 'BB', 'MACD', 'RSI', 'SMA', 'EMA'])

    bb_indicator = BollingerBands(data['Close'])
    data['bb_h'] = bb_indicator.bollinger_hband()
    data['bb_l'] = bb_indicator.bollinger_lband()
    bb = data[['Close', 'bb_h', 'bb_l']]

    macd = MACD(data['Close']).macd()
    rsi = RSIIndicator(data['Close']).rsi()
    sma = SMAIndicator(data['Close'], window=14).sma_indicator()
    ema = EMAIndicator(data['Close']).ema_indicator()

    if option == 'Close':
        st.write('Close Price')
        st.line_chart(data['Close'])
    elif option == 'BB':
        st.write('Bollinger Bands')
        st.line_chart(bb)
    elif option == 'MACD':
        st.write('Moving Average Convergence Divergence')
        st.line_chart(macd)
    elif option == 'RSI':
        st.write('Relative Strength Indicator')
        st.line_chart(rsi)
    elif option == 'SMA':
        st.write('Simple Moving Average')
        st.line_chart(sma)
    else:
        st.write('Exponential Moving Average')
        st.line_chart(ema)
    
    # Store data for email
    store_data_for_email(data)

def dataframe(data):
    st.header('Recent Data')
    st.dataframe(data.tail(10))
    
    # Store data for email
    store_data_for_email(data)

def predict(data):
    model = st.radio('Choose a model', ['LinearRegression', 'SVR', 'KNeighborsRegressor', 'XGBoostRegressor'])
    num = st.number_input('How many days forecast?', value=5)
    num = int(num)
    if st.button('Predict'):
        engine = get_model_engine(model)
        if engine:
            model_engine(engine, num, data)
    
    # Store data for email
    store_data_for_email(data)

def get_model_engine(model_name):
    if model_name == 'LinearRegression':
        return LinearRegression()
    elif model_name == 'SVR':
        return SVR()
    elif model_name == 'KNeighborsRegressor':
        return KNeighborsRegressor()
    elif model_name == 'XGBoostRegressor':
        return XGBRegressor(objective='reg:squarederror')  # Adjusted to use 'reg:squarederror'
    else:
        return None

def model_engine(model, num, data):
    data['returns'] = data['Close'].pct_change()
    data['log_returns'] = np.log(1 + data['returns'])
    data['bb_h'] = BollingerBands(data['Close']).bollinger_hband()
    data['bb_l'] = BollingerBands(data['Close']).bollinger_lband()
    data['macd'] = MACD(data['Close']).macd()
    data['rsi'] = RSIIndicator(data['Close']).rsi()
    data['sma'] = SMAIndicator(data['Close'], window=14).sma_indicator()
    data['ema'] = EMAIndicator(data['Close']).ema_indicator()

    features = ['Close', 'log_returns', 'bb_h', 'bb_l', 'macd', 'rsi', 'sma', 'ema']
    df = data[features]
    df['preds'] = data['Close'].shift(-num)
    df.dropna(inplace=True)

    x = df.drop(['preds'], axis=1).values
    y = df['preds'].values

    scaler = StandardScaler()
    x_scaled = scaler.fit_transform(x)

    x_train, x_test, y_train, y_test = train_test_split(x_scaled, y, test_size=0.2, random_state=7)
    model.fit(x_train, y_train)

    preds = model.predict(x_test)
    st.text(f'r2_score: {r2_score(y_test, preds)}\nMAE: {mean_absolute_error(y_test, preds)}')

    forecast_pred = model.predict(scaler.transform(x[-num:]))
    for day, prediction in enumerate(forecast_pred, start=1):
        st.text(f'Day {day}: {prediction}')
    
    # Store data for email
    store_data_for_email(data)

def sentiment_analysis(data, option):
    st.header('Sentiment Analysis')
    text = st.text_area('Enter the text for sentiment analysis (e.g., financial news headlines):')

    if st.button('Analyze Sentiment'):
        if text:
            analyzer = SentimentIntensityAnalyzer()
            sentiment = analyzer.polarity_scores(text)
            compound_score = sentiment['compound']
            email_sentiment_analysis(text, sentiment, compound_score, data, option)
            if compound_score > 0:
                st.success(f'Sentiment Analysis Result: Positive ({compound_score})')
            elif compound_score < 0:
                st.error(f'Sentiment Analysis Result: Negative ({compound_score})')
            else:
                st.info(f'Sentiment Analysis Result: Neutral ({compound_score})')

def email_sentiment_analysis(text, sentiment, compound_score, data, option):
    try:
        # Retrieve stored data
        stored_data = st.session_state.get('stored_data')

        stock_data_html = stored_data.to_html()
        subject = f"Sentiment Analysis and Stock Data for {option}"
        body = f"""
        <h3>Sentiment Analysis Result:</h3>
        <p>{text}</p>
        <p>Sentiment: {sentiment}</p>
        <p>Compound Score: {compound_score}</p>
        <h3>Recent Stock Data for {option}:</h3>
        {stock_data_html}
        """

        email_sender = "tanxstockmarketpredictions@gmail.com"
        password = "klah mrzk uyml shvf"

        msg = MIMEMultipart()
        msg['From'] = email_sender
        msg['To'] = st.session_state['email']
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_sender, password)
        server.sendmail(email_sender, st.session_state['email'], msg.as_string())
        server.quit()

        st.success('Email sent successfully! 📧')
    except Exception as e:
        st.error(f"Error sending email: {e}")

def generate_otp():
    return randint(1000, 9999)

def send_otp(email, otp):
    try:
        email_sender = "tanxstockmarketpredictions@gmail.com"
        password = "klah mrzk uyml shvf"

        msg = MIMEMultipart()
        msg['From'] = email_sender
        msg['To'] = email
        msg['Subject'] = "OTP for Login"

        body = f"Your OTP for login is: {otp}"
        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email_sender, password)
        server.sendmail(email_sender, email, msg.as_string())
        server.quit()

        st.success('OTP sent successfully! Check your email 📧')
    except Exception as e:
        st.error(f"Error sending OTP: {e}")

def otp_login():
    st.sidebar.title("Login with OTP")
    email = st.sidebar.text_input("Email")
    if st.sidebar.button("Send OTP"):
        try:
            validate_email(email)
            st.session_state['email'] = email
            otp = generate_otp()
            st.session_state['otp'] = otp
            send_otp(email, otp)
        except EmailNotValidError as e:
            st.sidebar.error(str(e))

    otp_input = st.sidebar.text_input("Enter OTP", type="password")
    if st.sidebar.button("Login"):
        if 'otp' in st.session_state and otp_input == str(st.session_state['otp']):
            st.session_state['logged_in'] = True
            log_user_login(email)  # Log user login to Excel
        else:
            st.sidebar.error("Invalid OTP")

def log_user_login(email):
    log_file = "user_logins.xlsx"
    current_time = dt.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(log_file):
        workbook = load_workbook(log_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Email", "Login Time"])

    sheet.append([email, current_time])
    workbook.save(log_file)

def store_data_for_email(data_to_store):
    st.session_state['stored_data'] = data_to_store

if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False

if not st.session_state['logged_in']:
    otp_login()
else:
    st.sidebar.info('Welcome to the Stock Price Prediction App. Choose your options below')

    option = st.sidebar.text_input('Enter a Stock Symbol', value='AAPL').upper()
    today = datetime.date.today()
    duration = st.sidebar.number_input('Enter the duration (days)', value=3000)
    before = today - datetime.timedelta(days=duration)
    start_date = st.sidebar.date_input('Start Date', value=before)
    end_date = st.sidebar.date_input('End date', value=today)

    if st.sidebar.button('Send'):
        if start_date < end_date:
            st.sidebar.success(f'Start date: `{start_date}`\n\nEnd date: `{end_date}`')
            data = download_data(option, start_date, end_date)
        else:
            st.sidebar.error('Error: End date must fall after start date')
    else:
        data = download_data(option, start_date, end_date)

    if data.empty:
        st.error("No data found. Please check the stock symbol and date range.")
    else:
        scaler = StandardScaler()

        option_selected = st.sidebar.selectbox('Make a choice', ['Visualize', 'Recent Data', 'Predict', 'Sentiment Analysis'])
        if option_selected == 'Visualize':
            tech_indicators(data)
        elif option_selected == 'Recent Data':
            dataframe(data)
        elif option_selected == 'Predict':
            predict(data)
        else:
            sentiment_analysis(data, option)
